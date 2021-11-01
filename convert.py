import os
import sys
import pandas as pd
import PySimpleGUI as sg
from datetime import date
import xlsxwriter
from reportlab.pdfgen.canvas import Canvas
from reportlab.lib.units import inch, cm
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
import openpyxl
from openpyxl.styles import Border, Side, Color
from openpyxl.cell import Cell
from openpyxl.styles import PatternFill
from typing import Optional
import subprocess

from win32com import client  # for excel to pdf

class makeFile:
    global_company = ''
    global_total1 = 0
    global_total2 = 0
    flag = 0
    point = 16
    d11 = ''
    d12 = ''
    c14 = ''
    n29 = ''
    d33 = ''
    f33 = ''
    g33 = ''
    h33 = ''
    navid_array = []
    start_array = []
    detail_array = []
    amount_array = []
    pay1_array = []
    pay2_array = []
    file_name = ''
    end_time = 1

    def makeEXCEl(self, data, today, excel_dir, sample_path, last_index, pdf_dir):
        if makeFile.flag == 0:
            count = 0
        elif makeFile.end_time == last_index:
            if data['15.会社名'] != makeFile.global_company:
                excel_data = openpyxl.load_workbook(sample_path)
                ws = excel_data.active

                ws['N2'] = '日付: ' + str(today)
                ws['D11'] = makeFile.d11
                ws['D12'] = makeFile.d12
                ws['C14'] = makeFile.c14
                ws['E11'] = ''

                ws['N29'] = makeFile.n29
                ws['D33'] = makeFile.d33
                ws['E33'] = makeFile.e33
                ws['F33'] = makeFile.f33
                ws['G33'] = makeFile.g33
                ws['H33'] = makeFile.h33

                medium = Side(border_style='medium', color="000000")
                left_border = Border(left=medium)
                right_border = Border(right=medium)
                for y in range(len(makeFile.navid_array)):
                    makeFile.point = 16 + y
                    if y > 8:
                        ws.insert_rows(makeFile.point)
                        ws['B' + str(makeFile.point) ].border = left_border
                        ws['O' + str(makeFile.point) ].border = right_border
                    ws['C' + str(makeFile.point) ] = makeFile.navid_array[y]                
                    ws['E' + str(makeFile.point) ] = makeFile.start_array[y]
                    ws['F' + str(makeFile.point) ] = makeFile.detail_array[y]
                    ws['K' + str(makeFile.point) ] = makeFile.amount_array[y]
                    if makeFile.pay1_array[y] < 0:
                        _pay_cell = ws['M' + str(makeFile.point)]
                        _pay_cell.font.color.rgb = 'FFFF0000'
                    else:
                        _pay_cell = ws['M' + str(makeFile.point)]
                        _pay_cell.font.color.rgb = 'FF000000'
                    if makeFile.pay2_array[y] < 0:
                        _pay_cell = ws['N' + str(makeFile.point)]
                        _pay_cell.font.color.rgb = 'FFFF0000'
                    else:
                        _pay_cell = ws['N' + str(makeFile.point)]
                        _pay_cell.font.color.rgb = 'FF000000'

                    mark_pay1 = makeFile.intWithCommas(makeFile.pay1_array[y])
                    mark_pay2 = makeFile.intWithCommas(makeFile.pay2_array[y])
                    ws['M' + str(makeFile.point) ] = str('¥') + str(mark_pay1)
                    ws['N' + str(makeFile.point) ] = str('¥') + str(mark_pay2)

                    makeFile.global_total1 += makeFile.pay1_array[y]
                    makeFile.global_total2 += makeFile.pay2_array[y]
                if makeFile.point < 25:
                    makeFile.point = 24
                if makeFile.global_total1 < 0:
                    _pay_cell = ws['M' + str(makeFile.point + 1)]
                    _pay_cell.font.color.rgb = 'FFFF0000'
                else:
                    _pay_cell = ws['M' + str(makeFile.point + 1)]
                    _pay_cell.font.color.rgb = 'FF000000'
                if makeFile.global_total2 < 0:
                    _pay_cell = ws['N' + str(makeFile.point + 1)]
                    _pay_cell.font.color.rgb = 'FFFF0000'
                else:
                    _pay_cell = ws['N' + str(makeFile.point + 1)]
                    _pay_cell.font.color.rgb = 'FF000000'
                mark_total1 = makeFile.intWithCommas(makeFile.global_total1)
                mark_total2 = makeFile.intWithCommas(makeFile.global_total2) 
                ws['M' + str(makeFile.point + 1)] =  str('¥') + str(mark_total1)
                ws['N' + str(makeFile.point + 1)] =  str('¥') + str(mark_total2)

                excel_path = os.path.join(excel_dir, makeFile.file_name)
                e_file_name = os.path.abspath(excel_path + '.xlsx')
                excel_data.save(e_file_name)

                # pdf create
                excel = client.Dispatch("Excel.Application")
                
                sheets = excel.Workbooks.Open(e_file_name)
                work_sheets = sheets.Worksheets[0]
                new_file_name = os.path.join(pdf_dir, makeFile.file_name)
                real_name = os.path.abspath(new_file_name + '.pdf')
                work_sheets.ExportAsFixedFormat(0, real_name)
                excel.Quit()
                
                makeFile.navid_array = []
                makeFile.start_array = []
                makeFile.detail_array = []
                makeFile.amount_array = []
                makeFile.pay1_array = []
                makeFile.pay2_array = []
                makeFile.global_total1 = 0
                makeFile.global_total2 = 0
                makeFile.point = 16
            makeFile.d11 = str(data['15.会社名']) + str('御中')
            makeFile.d12 = str(data['取引先口座管理'])
            makeFile.c14 = str(data['13.入庫日'])[0:10]
            makeFile.n29 = str(data['支払期限'])[0:10]
            makeFile.d33 = str(data['金融機関名（漢字） (from 取引先口座管理)'])
            makeFile.e33 = str(data['支店名(漢字)'])
            makeFile.f33 = str(data['科目'])
            makeFile.g33 = str(data['口座番号'])
            makeFile.h33 = str(data['口座名'])
            makeFile.navid_array.append(str(data['16.navid']))
            makeFile.start_array.append(str(data['14.入出庫作業開始日'])[0:10])
            makeFile.detail_array.append(str(data['17.明細区分']))
            makeFile.amount_array.append(str(data['19.確定数']))
            makeFile.pay1_array.append(data['M.総支払額(税抜)'])
            makeFile.pay2_array.append(data['M.総支払額(税込)'])
            makeFile.file_name = str(data['V支払グループ'])

            excel_data = openpyxl.load_workbook(sample_path)
            ws = excel_data.active
            ws['N2'] = '日付: ' + str(today)
            ws['D11'] = makeFile.d11
            ws['D12'] = makeFile.d12
            ws['C14'] = makeFile.c14
            ws['E11'] = ''

            ws['N29'] = makeFile.n29
            ws['D33'] = makeFile.d33
            ws['E33'] = makeFile.e33
            ws['F33'] = makeFile.f33
            ws['G33'] = makeFile.g33
            ws['H33'] = makeFile.h33

            medium = Side(border_style='medium', color="000000")
            left_border = Border(left=medium)
            right_border = Border(right=medium)
            for y in range(len(makeFile.navid_array)):
                makeFile.point = 16 + y
                if y > 8:
                    ws.insert_rows(makeFile.point)
                    ws['B' + str(makeFile.point) ].border = left_border
                    ws['O' + str(makeFile.point) ].border = right_border
                ws['C' + str(makeFile.point) ] = makeFile.navid_array[y]                
                ws['E' + str(makeFile.point) ] = makeFile.start_array[y]
                ws['F' + str(makeFile.point) ] = makeFile.detail_array[y]
                ws['K' + str(makeFile.point) ] = makeFile.amount_array[y]
                if makeFile.pay1_array[y] < 0:
                    _pay_cell = ws['M' + str(makeFile.point)]
                    _pay_cell.font.color.rgb = 'FFFF0000'
                else:
                    _pay_cell = ws['M' + str(makeFile.point)]
                    _pay_cell.font.color.rgb = 'FF000000'
                if makeFile.pay2_array[y] < 0:
                    _pay_cell = ws['N' + str(makeFile.point)]
                    _pay_cell.font.color.rgb = 'FFFF0000'
                else:
                    _pay_cell = ws['N' + str(makeFile.point)]
                    _pay_cell.font.color.rgb = 'FF000000'

                mark_pay1 = makeFile.intWithCommas(makeFile.pay1_array[y])
                mark_pay2 = makeFile.intWithCommas(makeFile.pay2_array[y])
                ws['M' + str(makeFile.point) ] = str('¥') + str(mark_pay1)
                ws['N' + str(makeFile.point) ] = str('¥') + str(mark_pay2)

                makeFile.global_total1 += makeFile.pay1_array[y]
                makeFile.global_total2 += makeFile.pay2_array[y]
            if makeFile.point < 25:
                makeFile.point = 24
            if makeFile.global_total1 < 0:
                _pay_cell = ws['M' + str(makeFile.point + 1)]
                _pay_cell.font.color.rgb = 'FFFF0000'
            else:
                _pay_cell = ws['M' + str(makeFile.point + 1)]
                _pay_cell.font.color.rgb = 'FF000000'
            if makeFile.global_total2 < 0:
                _pay_cell = ws['N' + str(makeFile.point + 1)]
                _pay_cell.font.color.rgb = 'FFFF0000'
            else:
                _pay_cell = ws['N' + str(makeFile.point + 1)]
                _pay_cell.font.color.rgb = 'FF000000'
            mark_total1 = makeFile.intWithCommas(makeFile.global_total1)
            mark_total2 = makeFile.intWithCommas(makeFile.global_total2) 
            ws['M' + str(makeFile.point + 1)] =  str('¥') + str(mark_total1)
            ws['N' + str(makeFile.point + 1)] =  str('¥') + str(mark_total2)

            excel_path = os.path.join(excel_dir, makeFile.file_name)
            e_file_name = os.path.abspath(excel_path + '.xlsx')
            excel_data.save(e_file_name)

            # pdf create
            excel = client.Dispatch("Excel.Application")
            
            sheets = excel.Workbooks.Open(e_file_name)
            work_sheets = sheets.Worksheets[0]
            new_file_name = os.path.join(pdf_dir, makeFile.file_name)
            real_name = os.path.abspath(new_file_name + '.pdf')
            work_sheets.ExportAsFixedFormat(0, real_name)
            excel.Quit()
        elif data['15.会社名'] != makeFile.global_company:
            excel_data = openpyxl.load_workbook(sample_path)
            ws = excel_data.active

            ws['N2'] = '日付: ' + str(today)
            ws['D11'] = makeFile.d11
            ws['D12'] = makeFile.d12
            ws['C14'] = makeFile.c14
            ws['E11'] = ''

            ws['N29'] = makeFile.n29
            ws['D33'] = makeFile.d33
            ws['E33'] = makeFile.e33
            ws['F33'] = makeFile.f33
            ws['G33'] = makeFile.g33
            ws['H33'] = makeFile.h33

            medium = Side(border_style='medium', color="000000")
            left_border = Border(left=medium)
            right_border = Border(right=medium)
            for y in range(len(makeFile.navid_array)):
                makeFile.point = 16 + y
                if y > 8:
                    ws.insert_rows(makeFile.point)
                    ws['B' + str(makeFile.point) ].border = left_border
                    ws['O' + str(makeFile.point) ].border = right_border
                ws['C' + str(makeFile.point) ] = makeFile.navid_array[y]                
                ws['E' + str(makeFile.point) ] = makeFile.start_array[y]
                ws['F' + str(makeFile.point) ] = makeFile.detail_array[y]
                ws['K' + str(makeFile.point) ] = makeFile.amount_array[y]
                if makeFile.pay1_array[y] < 0:
                    _pay_cell = ws['M' + str(makeFile.point)]
                    _pay_cell.font.color.rgb = 'FFFF0000'
                else:
                    _pay_cell = ws['M' + str(makeFile.point)]
                    _pay_cell.font.color.rgb = 'FF000000'
                if makeFile.pay2_array[y] < 0:
                    _pay_cell = ws['N' + str(makeFile.point)]
                    _pay_cell.font.color.rgb = 'FFFF0000'
                else:
                    _pay_cell = ws['N' + str(makeFile.point)]
                    _pay_cell.font.color.rgb = 'FF000000'

                mark_pay1 = makeFile.intWithCommas(makeFile.pay1_array[y])
                mark_pay2 = makeFile.intWithCommas(makeFile.pay2_array[y])
                ws['M' + str(makeFile.point) ] = str('¥') + str(mark_pay1)
                ws['N' + str(makeFile.point) ] = str('¥') + str(mark_pay2)

                makeFile.global_total1 += makeFile.pay1_array[y]
                makeFile.global_total2 += makeFile.pay2_array[y]
            if makeFile.point < 25:
                makeFile.point = 24
            if makeFile.global_total1 < 0:
                _pay_cell = ws['M' + str(makeFile.point + 1)]
                _pay_cell.font.color.rgb = 'FFFF0000'
            else:
                _pay_cell = ws['M' + str(makeFile.point + 1)]
                _pay_cell.font.color.rgb = 'FF000000'
            if makeFile.global_total2 < 0:
                _pay_cell = ws['N' + str(makeFile.point + 1)]
                _pay_cell.font.color.rgb = 'FFFF0000'
            else:
                _pay_cell = ws['N' + str(makeFile.point + 1)]
                _pay_cell.font.color.rgb = 'FF000000'
            mark_total1 = makeFile.intWithCommas(makeFile.global_total1)
            mark_total2 = makeFile.intWithCommas(makeFile.global_total2) 
            ws['M' + str(makeFile.point + 1)] =  str('¥') + str(mark_total1)
            ws['N' + str(makeFile.point + 1)] =  str('¥') + str(mark_total2)

            excel_path = os.path.join(excel_dir, makeFile.file_name)
            e_file_name = os.path.abspath(excel_path + '.xlsx')
            excel_data.save(e_file_name)

            # pdf create
            excel = client.Dispatch("Excel.Application")
            
            sheets = excel.Workbooks.Open(e_file_name)
            work_sheets = sheets.Worksheets[0]
            new_file_name = os.path.join(pdf_dir, makeFile.file_name)
            real_name = os.path.abspath(new_file_name + '.pdf')
            work_sheets.ExportAsFixedFormat(0, real_name)
            excel.Quit()
            
            makeFile.navid_array = []
            makeFile.start_array = []
            makeFile.detail_array = []
            makeFile.amount_array = []
            makeFile.pay1_array = []
            makeFile.pay2_array = []
            makeFile.global_total1 = 0
            makeFile.global_total2 = 0
            makeFile.point = 16
        else:
            count = 1

        makeFile.d11 = str(data['15.会社名']) + str('御中')
        makeFile.d12 = str(data['取引先口座管理'])
        makeFile.c14 = str(data['13.入庫日'])[0:10]
        makeFile.n29 = str(data['支払期限'])[0:10]
        makeFile.d33 = str(data['金融機関名（漢字） (from 取引先口座管理)'])
        makeFile.e33 = str(data['支店名(漢字)'])
        makeFile.f33 = str(data['科目'])
        makeFile.g33 = str(data['口座番号'])
        makeFile.h33 = str(data['口座名'])
        makeFile.navid_array.append(str(data['16.navid']))
        makeFile.start_array.append(str(data['14.入出庫作業開始日'])[0:10])
        makeFile.detail_array.append(str(data['17.明細区分']))
        makeFile.amount_array.append(str(data['19.確定数']))
        makeFile.pay1_array.append(data['M.総支払額(税抜)'])
        makeFile.pay2_array.append(data['M.総支払額(税込)'])
        makeFile.file_name = str(data['V支払グループ'])
        makeFile.global_company = data['15.会社名']
        makeFile.flag = 1
        makeFile.end_time = makeFile.end_time + 1

    def intWithCommas(self, x):
        if x < 0:
            return '-' + intWithCommas(-x)
        result = ''
        while x >= 1000:
            x, r = divmod(x, 1000)
            result = ",%03d%s" % (r, result)
        return "%d%s" % (x, result)

if __name__ == "__main__":
    makeFile = makeFile()
    sg.theme('Dark Blue 3')
    file_list_column = [
        [
            sg.Text('テンプレートExcelファイル', size=(25, 1)), sg.InputText(), sg.FileBrowse()
        ],
        [
            sg.Text('支払い明細書Excelファイル', size=(25, 1)), sg.InputText(), sg.FileBrowse()
        ],
        [
            sg.Text('自動明細書保存パス', size=(25, 1)), sg.InputText(), sg.FolderBrowse()
        ],
        [
            sg.ProgressBar(100, orientation='h', size=(54.4, 10), border_width=4, key='progbar',bar_color=['Red','Green'])
        ],
        [
            sg.Button('自動明細書生成')
        ]
    ]
    layout = [
        [
            sg.Column(file_list_column),
            sg.VSeperator(),
        ]
    ]
    window = sg.Window("Excelファイルから支払い明細書の自動作成", layout, margins=(50, 25))
    while True:
        event, values = window.read()
        if event == sg.WINDOW_CLOSED:
            break
        if event == '自動明細書生成':
            # テンプレートExcelファイルのパス取得
            sample_path = values[0]

            # 支払い明細書Excelファイルのパスとデータ取得
            payment_path = values[1]
            if payment_path == '':
                break
            pay_data = pd.read_excel(payment_path, index_col=1)

            # 今日の日付取得
            today = date.today()

            # フォルダ生成
            save_path = values[2]
            date_directory = today
            date_dir = os.path.join(save_path, str(date_directory))
            if os.path.exists(date_dir):
                count = 1
            else:
                os.mkdir(date_dir)

            pdf_dir = os.path.join(date_dir, 'pdf')
            if os.path.exists(pdf_dir):
                count = 1
            else:
                os.mkdir(pdf_dir)

            excel_dir = os.path.join(date_dir, 'excel')
            if os.path.exists(excel_dir):
                count = 1
            else:
                os.mkdir(excel_dir)

            val = 0
            temp_data = []
            for x in range(len(pay_data)):
                if pay_data.iloc[x]['13.入庫日'].value == -9223372036854775808:
                    count = 0
                else:
                    temp_data.append(pay_data.iloc[x])
            temp_data.sort(key=lambda x: (x.iloc[5], x.iloc[6]), reverse=False)

            for x in range(len(temp_data)):
                makeFile.makeEXCEl(temp_data[x], today, excel_dir, sample_path, len(temp_data), pdf_dir)
                val=val+100/(len(temp_data))
                window['progbar'].update_bar(val)
            break
    window.close()